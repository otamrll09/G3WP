import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# Monta o cabeçalho da planilha
cabecalho_excel = ['Software/Sistema', 'CVE', 'Descrição', 'Severidade CNA', 'Severidade NIST', 
                    'Referências para recomendações, soluções e ferramentas',
                    'Configurações de softwares afetadas', 'Data de publicação NVD', 'Link CVE']

# Recebe os dados obtidos via bot webscrapping
textos_site = ['texto1', 'texto2', 'texto3', 9, 2, 'texto4', 'texto7', 'texto8', 'texto9']

# Salva os dados em uma planilha do excel
dados_obtidos = pd.DataFrame(data=[textos_site],columns=cabecalho_excel, index=None)
dados_obtidos.to_excel('Teste_Data_Frame.xlsx', index=False)

# Abre o arquivo do excel para realizar a edição
wb = load_workbook(filename='Teste_Data_Frame.xlsx')
ws = wb['Sheet1']

# Varre as células nas colunas para configurar largura e alinhamento
for coluna in ws.columns:
    largura_maxima = 0
    indec_coluna = coluna[0].column_letter

    for celula in coluna:
        # Configuração de estilo de fonte
        if celula.row == 1:
            celula.font = Font(name='Arial', size=12, bold=True)
        else:
            celula.font = Font(name='Arial', size=12, bold=False)
        # Configuração de alinhamento/quebra de texto
        celula.alignment = Alignment(horizontal = 'center', vertical= 'center', wrap_text = True)
        try:
            if len(str(celula.value)) > largura_maxima:
                largura_maxima = len(str(celula.value))
        except:
            pass
    # Configuração de largura das colunas
    largura_ajustada = (largura_maxima+2) * 1.2
    if indec_coluna == 'F' or indec_coluna == 'C':
        ws.column_dimensions[indec_coluna].width = 66
    else:
        ws.column_dimensions[indec_coluna].width = largura_ajustada
        
# Comando para salvar a planilha
wb.save(filename='Teste_Data_Frame.xlsx')
