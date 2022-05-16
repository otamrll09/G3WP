def montaPlanilha(lst_felps):
    from cgitb import text
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    # Monta o cabeçalho da planilha
    cabecalho_excel = ['Software/Sistema', 'CVE', 'Descrição', 'Severidade', 
                        'Referências para recomendações, soluções e ferramentas',
                        'Configurações de softwares afetadas', 'Data de publicação NVD', 'Link CVE']

    # Salva os dados em uma planilha do excel
    dados_obtidos = pd.DataFrame(lst_felps,columns=cabecalho_excel, index=None)
    dados_obtidos.to_excel('VulnerabilidadesSolicitadas.xlsx', index=False)
    dados_gelso = pd.DataFrame(lst_felps,columns=cabecalho_excel, index=None)
    dados_gelso.to_dict()

    # Abre o arquivo do excel para realizar a edição
    wb = load_workbook(filename='VulnerabilidadesSolicitadas.xlsx')
    ws = wb['Sheet1']  

    # Varre as células nas colunas para configurar largura e alinhamento
    for coluna in ws.columns:
        largura_maxima = 0
        indec_coluna = coluna[0].column_letter

        for celula in coluna:
            # Limpa os caracteres
            for n in celula.value:
                if n == "'" or n == "[" or n == "]" or n == ",":
                    celula.value = celula.value.replace("'","")
                    celula.value = celula.value.replace("[","")
                    celula.value = celula.value.replace("]","")
                    if indec_coluna=="D" or indec_coluna=="E":
                        celula.value = celula.value.replace(",","\n")

            # Estiliza de acordo com o grau de severidade
            celula_severidade = 'D' + str(celula.row)
            if 'CRITICAL' in ws[celula_severidade].value:
                celula.fill = PatternFill(fill_type='solid',fgColor='FF0000')
            elif 'HIGH' in ws[celula_severidade].value:
                celula.fill = PatternFill(fill_type='solid',fgColor='FFFF00')
            else:
                celula.fill = PatternFill(fill_type='solid',fgColor='FFFFFF')

            # Configuração de estilo de fonte
            if celula.row == 1:
                celula.font = Font(name='Arial', size=12, bold=True)
            else:
                celula.font = Font(name='Arial', size=12, bold=False)

            # Configuração das bordas
            side = Side(style='thin', color='000000')
            celula.border = Border(left=side,right=side,top=side,bottom=side)

            # Configuração de alinhamento/quebra de texto
            celula.alignment = Alignment(horizontal = 'center', vertical= 'center', wrap_text = True)
            try:
                if len(str(celula.value)) > largura_maxima:
                    largura_maxima = len(str(celula.value))
            except:
                pass
        # Configuração de largura das colunas
        largura_ajustada = (largura_maxima+2) * 1.2
        if indec_coluna == 'F' or indec_coluna == 'C':
            ws.column_dimensions[indec_coluna].width = 66
        elif indec_coluna == 'E':
            ws.column_dimensions[indec_coluna].width = 100
        else:
            ws.column_dimensions[indec_coluna].width = largura_ajustada
            
    # Comando para salvar a planilha
    wb.save(filename='VulnerabilidadesSolicitadas.xlsx')

    return dados_gelso